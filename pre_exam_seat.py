import tkinter
import tkinter.ttk
from pathlib import Path
from tkinter import filedialog
import openpyxl


#tkinterのwidget(部品)
class Application(tkinter.Frame): #tkinterのフレームを継承
    def __init__(self,root): #rootを受け取る
        super().__init__(root, #基底クラスのイニシャライザを呼ぶ、rootを渡す
            width=700,height=350,
            borderwidth=4,relief='groove') #relief:境界線の種類
        self.root=root
        self.pack() #位置を設定して配置
        self.pack_propagate(0) #サイズ調整
        self.create_widgets()
    
    
    def create_widgets(self):
        '''
        quit_btn=tkinter.Button(self)
        quit_btn['text']='閉じる'
        quit_btn['command']=self.root.destroy
        quit_btn.pack(side='bottom') #bottomで一番下に設定
        '''

        self.text_box=tkinter.Entry(self, justify='center', width = 10, font = b)
        self.text_box.pack()

        chk = tkinter.Checkbutton(self, text='国文')
        chk.place(x=50, y=70)

        submit_btn=tkinter.Button(self, justify='center', width = 10, font = a)
        submit_btn['text']='座席番号表示'
        submit_btn['command']=self.save_data
        submit_btn.pack()

        self.message=tkinter.Message(self, aspect=500, justify='center', font=b)
        self.message.pack()


    
    def save_data(self):
        global retsu_number #関数内外でもアクセス可
        student_number_list = []
        text=self.text_box.get()
        text2 = text[1:9] #アスタリスクを抜く
        student_number = int(text2) #VLOOKUPを使いたいから数値に変換
        student_number_list.append(student_number) #excelに吐き出すように格納
        file_path = 'ddd.xlsx'
        wb=openpyxl.load_workbook(file_path,data_only=False) #Falseは書き込み可能


        '''「''」の間にその日の日付(月と日も記入)に変更。数字は半角で！'''
        ws = wb['3月27日']    #将来的にはシステム起動後にシートの自動生成、自動指定がしたい


        #ws=wb.worksheets[0] これよくわからんけど上の方がシート指定できて良さげ
        seat_number = ws['D'+str(retsu_number)] #座席番号取得
        ws['B'+str(retsu_number)].value=student_number #B列に生徒番号を入力
        wb.save(file_path) #上書き保存
        student_name = ws['C'+str(retsu_number)]
        #self.message['text']=student_name.value+'さんの席番号は '+str(seat_number.value)+' です。' #生徒氏名がVLOOKUPの式になってまう…属性valueでセルの値を取得
        self.message['text']='生徒番号'+text2+'の座席番号は\n ' +str(seat_number.value)+'番です。'
        self.text_box.delete(0, tkinter.END) #テキストボックス内を消す
        print(text2)
        retsu_number = retsu_number + 1 #列番号を下にずらし、入力値をずらす
        
    




root=tkinter.Tk()  #オブジェクト作成
root.title('過去問演習会座席管理表app')
root.geometry('800x400')
retsu_number = 3
a = ("Arial black", 15, "bold")
b = ("Arial black", 30, "bold")
app=Application(root=root)
app.mainloop()  #実行